"""
Générateur de planning annuel pour opérateurs (travail de nuit), avec équipes fixes,
week-ends décalés, vacances, formation, maladie et contrôle d'heures max.

Entrée : fichier Excel avec une feuille "Parametres" (deux colonnes : Nom, Valeur)
Champs attendus dans la colonne Nom :
- total_operators (int) : nombre total d'opérateurs
- operators_per_team (int) : nombre d'opérateurs par équipe
- working_days_per_week (int) : nombre de nuits travaillées par opérateur/équipe par semaine
- vacation_weeks_per_year (int) : nb de semaines de vacances/an par opérateur
- consecutive_vacation_weeks (int) : nb minimal de semaines consécutives par bloc de vacances
- training_days (int) : nb de jours de formation/an par opérateur
- sickness_days (int) : nb de jours de maladie/an par opérateur (réserve/estimé)
- max_working_hours (int/float) : nb max d'heures/an par opérateur
- hours_per_day (int/float) : nb d'heures par nuit travaillée
- Year (int) : année de planification (ex: 2026)

Sorties :
- planning.xlsx contenant :
  * "Schedule" : une ligne par opérateur et par date (Assignment ∈ {Work, Weekend, Vacation, Training, Sickness})
  * "Summary" : totaux par opérateur (jours et heures), contrôle des heures max
  * "Teams" : composition des équipes, tailles, paramètres lus

Hypothèses :
- Couverture minimale : 1 équipe opérationnelle chaque nuit. Le schéma hebdomadaire fixe
  (mêmes jours de repos pour tous les membres d'une équipe) est décalé entre équipes pour couvrir 7 nuits.
- Si le nombre d'équipes disponibles < nombre minimal requis pour couvrir 7 nuits avec le working_days_per_week,
  le script produit un avertissement et construit tout de même le meilleur décalage possible.
- Les vacances/formation/maladie sont lissées dans l'année pour éviter les pics, mais sans logique avancée
  de limite de congés simultanés au sein d'une équipe (facile à ajouter si besoin).

Utilisation :
1) Placez ce script à côté de votre fichier d'entrée, par ex. Parametres.xlsx
2) Lancez : python plan_nuit.py --input Parametres.xlsx --output planning.xlsx
3) Ouvrez planning.xlsx

"""

from __future__ import annotations
import argparse
import math
from dataclasses import dataclass
from typing import List, Dict, Tuple
import pandas as pd
import numpy as np
from datetime import date, datetime, timedelta

# Boîtes de dialogue (sélection des fichiers)
try:
    from tkinter import Tk, filedialog  # type: ignore
except Exception:  # environnement sans interface graphique
    Tk = None
    filedialog = None


# ---------------------- Utilitaires calendrier ----------------------

def daterange(start: date, end: date):
    cur = start
    while cur <= end:
        yield cur
        cur = cur + timedelta(days=1)


def iso_week_start(d: date) -> date:
    # Lundi de la semaine ISO
    return d - timedelta(days=(d.weekday()))


# ---------------------- Lecture paramètres ----------------------

def read_params(path: str) -> Dict[str, float]:
    df = pd.read_excel(path, sheet_name="Parameters_nights")
    # tolérer variantes de casse/espaces
    df.columns = [str(c).strip() for c in df.columns]
    nom_col    = 'Nom'    if 'Nom'    in df.columns else 'Name'
    valeur_col = 'Valeur' if 'Valeur' in df.columns else 'Value'
    if nom_col not in df.columns or valeur_col not in df.columns:
        raise ValueError("The sheet Parameters_nights must contain 'Nom'/'Name' and 'Valeur'/'Value'.")
    params = (
        df.assign(**{nom_col: lambda x: x[nom_col].astype(str).str.strip()})
          .set_index(nom_col)[valeur_col].to_dict()
    )
    # conversions typées
    int_keys = [
        "total_operators","operators_per_team","working_days_per_week",
        "vacation_weeks_per_year","consecutive_vacation_weeks",
        "training_days","sickness_days","Year"
    ]
    float_keys = ["max_working_hours","hours_per_day"]

    out = {}
    for k in int_keys:
        if k not in params:
            raise ValueError(f"Missing paramater: {k}")
        out[k] = int(params[k])
    for k in float_keys:
        if k not in params:
            raise ValueError(f"Missing paramater: {k}")
        out[k] = float(params[k])

    # Optional for backward compatibility. Most night schedules are 7-day weeks.
    out["days_per_week"] = int(params.get("days_per_week", 7))
    out["use_available_capacity"] = str(params.get("use_available_capacity", "Yes")).strip().lower() in {
        "yes", "y", "true", "1", "oui"
    }
    out["vacation_preferred_start_month"] = int(params.get("vacation_preferred_start_month", 6))
    out["vacation_preferred_end_month"] = int(params.get("vacation_preferred_end_month", 9))

    validate_params(out)
    return out


def validate_params(P: Dict[str, float]) -> None:
    if P["total_operators"] <= 0:
        raise ValueError("total_operators must be > 0")
    if P["operators_per_team"] <= 0:
        raise ValueError("operators_per_team must be > 0")
    if P["days_per_week"] <= 0:
        raise ValueError("days_per_week must be > 0")
    if P["working_days_per_week"] <= 0:
        raise ValueError("working_days_per_week must be > 0")
    if P["working_days_per_week"] > P["days_per_week"]:
        raise ValueError("working_days_per_week cannot be greater than days_per_week")
    if P["consecutive_vacation_weeks"] <= 0:
        raise ValueError("consecutive_vacation_weeks must be > 0")
    if P["vacation_weeks_per_year"] < 0:
        raise ValueError("vacation_weeks_per_year must be >= 0")
    if P["training_days"] < 0:
        raise ValueError("training_days must be >= 0")
    if P["sickness_days"] < 0:
        raise ValueError("sickness_days must be >= 0")
    if P["hours_per_day"] <= 0:
        raise ValueError("hours_per_day must be > 0")
    if P["max_working_hours"] <= 0:
        raise ValueError("max_working_hours must be > 0")
    if not 1 <= P["vacation_preferred_start_month"] <= 12:
        raise ValueError("vacation_preferred_start_month must be between 1 and 12")
    if not 1 <= P["vacation_preferred_end_month"] <= 12:
        raise ValueError("vacation_preferred_end_month must be between 1 and 12")
    if P["vacation_preferred_start_month"] > P["vacation_preferred_end_month"]:
        raise ValueError("vacation_preferred_start_month must be <= vacation_preferred_end_month")


# ---------------------- Construction des équipes ----------------------

def build_teams(total_ops: int, team_size: int) -> Dict[str, List[str]]:
    teams: Dict[str, List[str]] = {}
    if team_size <= 0:
        raise ValueError("operators_per_team must be > 0")
    n_teams = math.ceil(total_ops / team_size)
    op_ids = [f"OP{str(i+1).zfill(3)}" for i in range(total_ops)]
    idx = 0
    for t in range(1, n_teams+1):
        team_name = f"Team_{t:02d}"
        teams[team_name] = []
        for _ in range(team_size):
            if idx < total_ops:
                teams[team_name].append(op_ids[idx])
                idx += 1
    return teams


def min_teams_required(working_days_per_week: int) -> int:
    if working_days_per_week <= 0:
        raise ValueError("working_days_per_week must be > 0")
    return math.ceil(7 / working_days_per_week)


def min_operators_required(days_per_week: int, required_per_day: int, working_days_per_week: int) -> int:
    return math.ceil((days_per_week * required_per_day) / working_days_per_week)


def consecutive_days(start: int, length: int, days_per_week: int) -> List[int]:
    return [(start + i) % days_per_week for i in range(length)]


def rest_start_order(rest_days: int, days_per_week: int) -> List[int]:
    """Start with a classic weekend block, then spread rest blocks through the week."""
    if rest_days <= 0:
        return []
    base_start = (days_per_week - rest_days) % days_per_week
    starts = []
    for i in range(days_per_week):
        starts.append((base_start + i * rest_days) % days_per_week)
    for i in range(days_per_week):
        if i not in starts:
            starts.append(i)
    return starts


def build_operator_rest_patterns(
    operators: List[str],
    days_per_week: int,
    working_days_per_week: int,
    required_per_day: int,
) -> tuple[Dict[str, List[int]], List[str]]:
    """Assign each operator a consecutive rest block while spreading coverage."""
    rest_days = days_per_week - working_days_per_week
    warnings: List[str] = []
    if rest_days == 0:
        return {op: [] for op in operators}, warnings

    max_resting_per_day = len(operators) - required_per_day
    if max_resting_per_day < 0:
        warnings.append(
            f"WARNING: total_operators ({len(operators)}) is lower than operators_per_team "
            f"({required_per_day}); daily coverage is impossible."
        )
        max_resting_per_day = 0

    counts = {d: 0 for d in range(days_per_week)}
    starts = rest_start_order(rest_days, days_per_week)
    patterns: Dict[str, List[int]] = {}

    for op in operators:
        candidates = []
        for order_idx, start in enumerate(starts):
            block = consecutive_days(start, rest_days, days_per_week)
            overflow = sum(max(0, counts[d] + 1 - max_resting_per_day) for d in block)
            load = sum(counts[d] for d in block)
            peak = max(counts[d] + 1 for d in block)
            candidates.append((overflow, load, peak, order_idx, start, block))
        _, _, _, _, chosen_start, chosen_block = min(candidates, key=lambda x: (x[0], x[1], x[2], x[3]))
        patterns[op] = chosen_block
        for d in chosen_block:
            counts[d] += 1

    weekly_coverage = {d: len(operators) - counts[d] for d in range(days_per_week)}
    weak_days = [d for d, available in weekly_coverage.items() if available < required_per_day]
    if weak_days:
        labels = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        weak_labels = [labels[d] if d < len(labels) else str(d) for d in weak_days]
        warnings.append(
            "WARNING: the theoretical weekly rest pattern cannot cover "
            f"{required_per_day} operators on: {', '.join(weak_labels)}."
        )

    return patterns, warnings


def weekly_rest_pattern_for_team(team_index: int, w: int) -> List[int]:
    """Retourne la liste des jours de repos (0=Lundi .. 6=Dimanche) pour une équipe.
    On crée un bloc de w jours travaillés consécutifs et (7-w) jours de repos.
    Chaque équipe décale sa fenêtre de repos pour répartir la couverture.
    """
    work_block = w
    rest_block = 7 - w
    # Par défaut : repos consécutifs (rest_block), par ex. Sam-Dim pour w=5
    base_rest_start = (5 - rest_block) % 7  # heuristique : repos vers fin de semaine
    # Décalage par index d'équipe pour étaler les repos
    shift = team_index % 7
    start = (base_rest_start + shift) % 7
    rest_days = [ (start + i) % 7 for i in range(rest_block) ]
    return rest_days


# ---------------------- Affectations spéciales (vacances/formation/maladie) ----------------------

def plan_vacations_for_operator(year: int, total_weeks: int, vac_weeks: int, min_block: int, seed: int) -> List[Tuple[date, date]]:
    """Renvoie des intervalles [start, end] (dates incluses) sur des semaines ISO complètes.
    Répartit en blocs de min_block semaines (dernier bloc peut être plus court si nécessaire).
    """
    rng = np.random.default_rng(seed)
    blocks = []
    if vac_weeks <= 0:
        return blocks
    # Découper en blocs
    full_blocks, remainder = divmod(vac_weeks, min_block)
    sizes = [min_block]*full_blocks + ([remainder] if remainder else [])
    # Générer des semaines candidates (1..total_weeks), éviter les collisions
    chosen = set()
    for sz in sizes:
        # tenter jusqu'à 200 essais pour cas pathologiques
        for _ in range(200):
            wk = int(rng.integers(1, total_weeks - sz + 2))  # semaine de début
            if all((wk+i) not in chosen for i in range(sz)):
                for i in range(sz):
                    chosen.add(wk+i)
                # convertir semaine -> dates
                start = iso_week_start(date.fromisocalendar(year, wk, 1))
                end = start + timedelta(days=7*sz - 1)
                blocks.append((start, end))
                break
    # trier par date
    blocks.sort(key=lambda ab: ab[0])
    return blocks


def vacation_block_sizes(vac_weeks: int, min_block: int) -> List[int]:
    if vac_weeks <= 0:
        return []
    if vac_weeks <= min_block:
        return [vac_weeks]
    full_blocks, remainder = divmod(vac_weeks, min_block)
    sizes = [min_block] * full_blocks
    if remainder:
        sizes[0] += remainder
    return sizes


def month_window(year: int, start_month: int, end_month: int) -> Tuple[date, date]:
    start = date(year, start_month, 1)
    if end_month == 12:
        end = date(year, 12, 31)
    else:
        end = date(year, end_month + 1, 1) - timedelta(days=1)
    return start, end


def plan_vacations_for_operator(
    year: int,
    total_weeks: int,
    vac_weeks: int,
    min_block: int,
    seed: int,
    preferred_start_month: int = 6,
    preferred_end_month: int = 9,
) -> List[Tuple[date, date]]:
    """Return vacation blocks, preferring full blocks inside the configured peak months."""
    rng = np.random.default_rng(seed)
    blocks: List[Tuple[date, date]] = []
    if vac_weeks <= 0:
        return blocks

    sizes = vacation_block_sizes(vac_weeks, min_block)
    chosen = set()
    preferred_start, preferred_end = month_window(year, preferred_start_month, preferred_end_month)

    def candidate_weeks(sz: int, preferred_only: bool) -> List[int]:
        weeks = []
        for wk in range(1, total_weeks - sz + 2):
            start = iso_week_start(date.fromisocalendar(year, wk, 1))
            end = start + timedelta(days=7 * sz - 1)
            if preferred_only and not (start >= preferred_start and end <= preferred_end):
                continue
            if all((wk + i) not in chosen for i in range(sz)):
                weeks.append(wk)
        return weeks

    for sz in sizes:
        candidates = candidate_weeks(sz, preferred_only=True)
        if not candidates:
            candidates = candidate_weeks(sz, preferred_only=False)
        if not candidates:
            continue

        wk = int(candidates[int(rng.integers(0, len(candidates)))])
        for i in range(sz):
            chosen.add(wk + i)
        start = iso_week_start(date.fromisocalendar(year, wk, 1))
        end = start + timedelta(days=7 * sz - 1)
        blocks.append((start, end))

    blocks.sort(key=lambda ab: ab[0])
    return blocks


def pick_scattered_days(year_dates: List[date], k: int, avoid_weekends: bool, seed: int) -> List[date]:
    rng = np.random.default_rng(seed)
    pool = [d for d in year_dates if not (avoid_weekends and d.weekday()>=5)]
    if k <= 0 or len(pool) == 0:
        return []
    if k >= len(pool):
        return sorted(pool)
    # échantillonnage stratifié par mois pour lisser
    by_month: Dict[int, List[date]] = {}
    for d in pool:
        by_month.setdefault(d.month, []).append(d)
    picks = []
    remaining = k
    months = list(by_month.keys())
    i = 0
    while remaining > 0 and months:
        m = months[i % len(months)]
        if by_month[m]:
            idx = int(rng.integers(0, len(by_month[m])))
            picks.append(by_month[m].pop(idx))
            remaining -= 1
            if not by_month[m]:
                months.remove(m)
                i -= 1
        i += 1
    return sorted(picks)


def pick_scattered_days_from_pool(pool: List[date], k: int, seed: int) -> List[date]:
    """Pick days from an already eligible pool, spread by month where possible."""
    rng = np.random.default_rng(seed)
    pool = sorted(set(pool))
    if k <= 0 or not pool:
        return []
    if k >= len(pool):
        return pool

    by_month: Dict[int, List[date]] = {}
    for d in pool:
        by_month.setdefault(d.month, []).append(d)

    picks: List[date] = []
    remaining = k
    months = sorted(by_month.keys())
    i = 0
    while remaining > 0 and months:
        m = months[i % len(months)]
        idx = int(rng.integers(0, len(by_month[m])))
        picks.append(by_month[m].pop(idx))
        remaining -= 1
        if not by_month[m]:
            months.remove(m)
            i -= 1
        i += 1
    return sorted(picks)

def ensure_min_coverage(schedule: pd.DataFrame, P: dict) -> tuple[pd.DataFrame, list]:
    """
    Garantit ≥1 opérateur par nuit.
    Si une nuit n'a aucun 'Work', on convertit un 'Weekend' éligible en 'Work'
    (pas Vacation/Training/Sickness, et sans dépasser max_working_hours).
    On choisit l'opérateur avec le moins d'heures déjà travaillées (priorité à l'équilibre secondaire).
    Retourne (schedule_modifié, dates_non_couvertes_restantes)
    """
    hpd = float(P["hours_per_day"])
    cap = float(P["max_working_hours"])

    # s'assurer du dtype
    schedule = schedule.copy()
    schedule["Date"] = pd.to_datetime(schedule["Date"], errors="raise")

    # heures déjà comptées par opérateur
    def hours_worked(df):
        return (df[df["Assignment"]=="Work"]
                .groupby("Operator")["Hours"]
                .sum()
                .reindex(df["Operator"].unique(), fill_value=0))

    # boucle sur dates
    uncovered_left = []
    for d in sorted(schedule["Date"].unique()):
        on_duty = (schedule["Assignment"].eq("Work") & schedule["Date"].eq(d)).sum()
        if on_duty >= 1:
            continue

        # candidats = opérateurs en Weekend ce jour-là (donc pas d'absence)
        cand = schedule.index[schedule["Date"].eq(d) & schedule["Assignment"].eq("Weekend")]
        if len(cand) == 0:
            uncovered_left.append(pd.Timestamp(d))
            continue

        # calc heures actuelles par opérateur
        hw = hours_worked(schedule)

        # filtrer ceux qui respectent le plafond si on ajoute hpd
        elig = []
        for idx in cand:
            op = schedule.at[idx, "Operator"]
            if float(hw.get(op, 0)) + hpd <= cap:
                elig.append((idx, float(hw.get(op, 0))))
        if not elig:
            uncovered_left.append(pd.Timestamp(d))
            continue

        # choisir le moins chargé (heures), puis par Team/Operator pour stabilité
        elig_sorted = sorted(
            ((idx, hrs,
              schedule.at[idx, "Team"], schedule.at[idx, "Operator"]) for idx, hrs in elig),
            key=lambda t: (t[1], t[2], t[3])
        )
        pick_idx = elig_sorted[0][0]

        # bascule Weekend -> Work
        schedule.at[pick_idx, "Assignment"] = "Work"
        schedule.at[pick_idx, "Hours"] = hpd

    return schedule, uncovered_left


# ---------------------- Génération du planning ----------------------

def generate_schedule(params_path: str, output_path: str):
    P = read_params(params_path)

    total_ops = P["total_operators"]
    team_size = P["operators_per_team"]
    w = P["working_days_per_week"]
    vac_weeks = P["vacation_weeks_per_year"]
    vac_min_block = P["consecutive_vacation_weeks"]
    training_days = P["training_days"]
    sickness_days = P["sickness_days"]
    max_hours = P["max_working_hours"]
    hpd = P["hours_per_day"]
    year = P["Year"]

    # Dates de l'année
    start = date(year, 1, 1)
    end = date(year, 12, 31)
    dates = [d for d in daterange(start, end)]
    total_weeks = date(year, 12, 28).isocalendar().week  # au moins 52, parfois 53

    # Équipes
    teams = build_teams(total_ops, team_size)
    team_names = list(teams.keys())
    n_teams = len(team_names)
    needed = min_teams_required(w)

    warnings = []
    if n_teams < needed:
        warnings.append(
            f"WARNING: {n_teams} teams available, but at least {needed} are required to cover 7 nights with w={w}.\n"
            "The script will still spread rest days to achieve the most even coverage possible."
        )

    # Modèle hebdo par équipe : jours de repos (0=Lun..6=Dim)
    team_rest_days: Dict[str, List[int]] = {}
    for i, tn in enumerate(team_names):
        team_rest_days[tn] = weekly_rest_pattern_for_team(i, w)

    # Table Schedule (pleine grille opérateurs x dates)
    rows = []

    # Préparer absences par opérateur
    all_seed = 2025 + year  # base pour reproductibilité
    op_to_vac_blocks: Dict[str, List[Tuple[date, date]]] = {}
    op_to_training_days: Dict[str, List[date]] = {}
    op_to_sickness_days: Dict[str, List[date]] = {}

    for t_idx, tn in enumerate(team_names):
        for j, op in enumerate(teams[tn]):
            seed = all_seed + (t_idx+1)*1000 + (j+1)
            vac_blocks = plan_vacations_for_operator(year, total_weeks, vac_weeks, vac_min_block, seed)
            op_to_vac_blocks[op] = vac_blocks
            # jours de formation/maladie : éviter weekend
            op_to_training_days[op] = pick_scattered_days(dates, training_days, avoid_weekends=True, seed=seed+7)
            op_to_sickness_days[op] = pick_scattered_days(dates, sickness_days, avoid_weekends=True, seed=seed+17)

    # Affectation initiale Work/Weekend par équipe selon pattern hebdo
    date_weekday = {d: d.weekday() for d in dates}

    for tn in team_names:
        rest_wd = set(team_rest_days[tn])
        members = teams[tn]
        for op in members:
            vac_blocks = op_to_vac_blocks[op]
            vac_mask = set()
            for a, b in vac_blocks:
                for d in daterange(a, b):
                    vac_mask.add(d)
            training_set = set(op_to_training_days[op])
            sickness_set = set(op_to_sickness_days[op])

            for d in dates:
                wd = date_weekday[d]
                # Base : Weekend si jour de repos d'équipe, sinon Work
                assignment = "Weekend" if wd in rest_wd else "Work"
                # Absences personnelles priment sur Work
                if d in vac_mask:
                    assignment = "Vacation"
                elif d in training_set:
                    assignment = "Training"
                elif d in sickness_set:
                    assignment = "Sickness"
                rows.append({
                    "Date": d,
                    "Weekday": ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"][wd],
                    "Team": tn,
                    "Operator": op,
                    "Assignment": assignment,
                    "Hours": hpd if assignment == "Work" else 0.0,
                })

    schedule = pd.DataFrame(rows)

    # Contrôle heures max -> réduction si dépassement
    def enforce_hours_cap(df: pd.DataFrame) -> pd.DataFrame:
        out = df.copy()
        work_mask = out["Assignment"] == "Work"
        hours_by_op = out.loc[work_mask].groupby("Operator")["Hours"].sum().reindex(out["Operator"].unique(), fill_value=0)
        over = hours_by_op - max_hours
        over = over[over > 0]
        if over.empty:
            return out
        # Pour chaque opérateur en dépassement, convertir en Weekend des quarts (du plus tard vers le plus tôt)
        by_date_desc = out.sort_values("Date", ascending=False)
        for op, extra in over.items():
            need_nights_to_drop = int(math.ceil(extra / hpd))
            idxs = by_date_desc.index[(by_date_desc["Operator"]==op) & (by_date_desc["Assignment"]=="Work")]
            to_flip = list(idxs)[:need_nights_to_drop]
            out.loc[to_flip, "Assignment"] = "Weekend"
            out.loc[to_flip, "Hours"] = 0.0
        return out

    schedule = enforce_hours_cap(schedule)

    # Priorité 1 : assurer la couverture minimale
    schedule, uncovered_left = ensure_min_coverage(schedule, P)

    # (optionnel) message console si des nuits restent impossibles à couvrir
    if uncovered_left:
        print(f"[Alerte] {len(uncovered_left)} night(s) remain uncovered (hours constraint or absences).")

    # S'assurer du dtype
    schedule["Date"] = pd.to_datetime(schedule["Date"], errors="raise")
    schedule["WeekdayIdx"] = schedule["Date"].dt.weekday

    # Taille d'équipe (pour normaliser)
    team_sizes = pd.Series({t: len(members) for t, members in teams.items()}, name="Team_size")
    team_sizes = team_sizes.rename_axis("Team").reset_index()

    # --- A) Nuits "Work" par (Date, Team) => Ops (nombre d'opérateurs)
    per_team_date_ops = (schedule.query('Assignment == "Work"')
                                .groupby(["Date", "Team"]).size()
                                .rename("Ops").reset_index())
    per_team_date_ops = per_team_date_ops.merge(team_sizes, on="Team", how="left")
    per_team_date_ops["WeekdayIdx"] = per_team_date_ops["Date"].dt.weekday

    # --- B) Normalisation par opérateur (Ops / Team_size) -> "operator-nights"
    per_team_date_ops["Ops_norm"] = per_team_date_ops["Ops"] / per_team_date_ops["Team_size"]

    # Pivot par jour de semaine (somme annuelle)
    per_team_weekday_ops_norm = (per_team_date_ops
        .pivot_table(index="Team", columns="WeekdayIdx", values="Ops_norm",
                    aggfunc="sum", fill_value=0)
        .rename(columns={0:"Mon",1:"Tue",2:"Wed",3:"Thu",4:"Fri",5:"Sat",6:"Sun"})
        .reset_index())

    # --- C) Totaux normalisés annuels (utile pour comparer les équipes directement)
    per_team_ops_total = (schedule.query('Assignment == "Work"')
                                .groupby("Team").size()
                                .rename("Ops_total").reset_index())
    per_team_norm = (per_team_ops_total.merge(team_sizes, on="Team", how="left")
                                    .assign(Nights_per_operator=lambda x: (x["Ops_total"] / x["Team_size"]).round(2)))

    # --- D) (Optionnel) Présence binaire et "équipe complète" (pour l'analyse)
    per_team_date_ops["Team_present"] = per_team_date_ops["Ops"] >= 1
    present_pivot = (per_team_date_ops
        .pivot_table(index="Team", columns="WeekdayIdx", values="Team_present",
                    aggfunc="sum", fill_value=0)
        .rename(columns={0:"Mon",1:"Tue",2:"Wed",3:"Thu",4:"Fri",5:"Sat",6:"Sun"})
        .reset_index())

    per_team_date_ops["Team_full"] = per_team_date_ops["Ops"] >= P["operators_per_team"]
    full_pivot = (per_team_date_ops
        .pivot_table(index="Team", columns="WeekdayIdx", values="Team_full",
                    aggfunc="sum", fill_value=0)
        .rename(columns={0:"Mon",1:"Tue",2:"Wed",3:"Thu",4:"Fri",5:"Sat",6:"Sun"})
        .reset_index())

    # ---------------------- Load balance counters & reports ----------------------
    # Compteurs (nombres de nuits "Work")
    per_op = (schedule.query('Assignment == "Work"')
                    .groupby(["Operator", "Team"]).size()
                    .rename("Work_nights").reset_index())

    per_team = (schedule.query('Assignment == "Work"')
                        .groupby("Team").size()
                        .rename("Work_nights").reset_index())

    # Assurer un dtype datetime64[ns] pour utiliser .dt
    schedule["Date"] = pd.to_datetime(schedule["Date"], errors="raise")

    # Par jour de semaine (0=Mon..6=Sun)
    schedule["WeekdayIdx"] = schedule["Date"].dt.weekday
    per_team_weekday = (schedule.query('Assignment == "Work"')
                            .groupby(["Team", "WeekdayIdx"]).size()
                            .rename("Work_nights").reset_index())

    # Agrégats utiles
    import numpy as np
    per_op_stats = per_op.groupby("Team")["Work_nights"].agg(
        mean="mean", std="std", min="min", max="max", sum="sum", count="count"
    ).reset_index().rename(columns={"count": "Operators_in_team"})

    per_team_weekday_pivot = per_team_weekday.pivot_table(
        index="Team", columns="WeekdayIdx", values="Work_nights", aggfunc="sum", fill_value=0
    ).reset_index()
    per_team_weekday_pivot.columns = ["Team","Mon","Tue","Wed","Thu","Fri","Sat","Sun"]

    # Version simple de 'cover' : nb d'opérateurs actifs par date (tu l'as déjà)
    # -> On ajoute un indicateur d'écart par jour de semaine (utile pour voir la régularité)
    weekday_load = (schedule.query('Assignment == "Work"')
                            .groupby("WeekdayIdx").size()
                            .rename("Ops_on_duty_sum").reset_index())
    weekday_load["Label"] = weekday_load["WeekdayIdx"].map({
        0: "Mon", 1: "Tue", 2: "Wed", 3: "Thu", 4: "Fri", 5: "Sat", 6: "Sun"
    })
    total_weeks = len(schedule["Date"].dt.isocalendar().week.unique())
    weekday_load["Avg_ops_per_night"] = (weekday_load["Ops_on_duty_sum"] / total_weeks).round(2)

    # Facultatif : petite alerte console si gros déséquilibre intra-équipe
    imbalance = per_op_stats["std"].fillna(0)
    if (imbalance > 2).any():  # seuil arbitraire = 2 nuits d'écart-type
        print("[Alerte] Imbalance detected in one or more teams (std dev > 2 nights).")



    # Couverture par nuit (nb d'opérateurs en Work)
    cover = (schedule[schedule["Assignment"]=="Work"].groupby("Date").size()
             .reindex(sorted(schedule["Date"].unique()), fill_value=0)
             .rename("Operators_on_duty").reset_index())

    # -> Ajouter le jour de la semaine
    cover["Date"] = pd.to_datetime(cover["Date"], errors="raise")
    cover["Weekday"] = cover["Date"].dt.dayofweek.map({
        0:"Mon",1:"Tue",2:"Wed",3:"Thu",4:"Fri",5:"Sat",6:"Sun"
    })
    # Réordonner les colonnes : Date | Weekday | Operators_on_duty
    cover = cover[["Date", "Weekday", "Operators_on_duty"]]
    
    # Résumés par opérateur
    piv = schedule.pivot_table(index=["Operator","Team"], columns="Assignment", values="Date", aggfunc="count", fill_value=0)
    for col in ["Work","Weekend","Vacation","Training","Sickness"]:
        if col not in piv.columns:
            piv[col] = 0
    piv = piv.reset_index()
    hours_work = (schedule[schedule["Assignment"]=="Work"].groupby(["Operator","Team"])['Hours'].sum()
                  .reindex(piv.set_index(["Operator","Team"]).index, fill_value=0).values)
    piv["Hours_worked"] = hours_work
    piv["Hours_cap"] = max_hours
    piv["Cap_margin"] = piv["Hours_cap"] - piv["Hours_worked"]

    # Feuille Teams/Params
    team_rows = []
    for tn, members in teams.items():
        team_rows.append({"Team": tn, "Team_size": len(members), "Members": ", ".join(members)})
    teams_df = pd.DataFrame(team_rows)

    params_df = pd.DataFrame([{"Param": k, "Value": v} for k, v in P.items()])

    # Écriture Excel
    with pd.ExcelWriter(output_path, engine="xlsxwriter", datetime_format="yyyy-mm-dd") as writer:
        schedule.sort_values(["Date","Team","Operator"]).to_excel(writer, sheet_name="Schedule", index=False)
        piv.sort_values(["Team","Operator"]).to_excel(writer, sheet_name="Summary", index=False)
        teams_df.to_excel(writer, sheet_name="Teams", index=False)
        cover.to_excel(writer, sheet_name="Coverage", index=False)
        params_df.to_excel(writer, sheet_name="Parameters", index=False)
        # Mise en forme simple
        # >>> NEW: Load_Balance (plusieurs tables)
        # Load_Balance - départ
        per_op_sorted = per_op.sort_values(["Team","Operator"])
        per_op_sorted.to_excel(writer, sheet_name="Load_Balance", startrow=0, index=False)
        ws = writer.sheets["Load_Balance"]
        ws.write(0, 0, "Per-Operator Work Nights")

        row = len(per_op_sorted) + 3
        ws.write(row, 0, "Per-Team Work Nights")
        per_team_sorted = per_team.sort_values("Team")
        per_team_sorted.to_excel(writer, sheet_name="Load_Balance", startrow=row+1, index=False)

        row = row + len(per_team_sorted) + 4
        ws.write(row, 0, "Per-Team by Weekday (Work Nights)")
        per_team_weekday_pivot.to_excel(writer, sheet_name="Load_Balance", startrow=row+1, index=False)

        row = row + len(per_team_weekday_pivot) + 4
        ws.write(row, 0, "Intra-Team Stats (Work Nights per Operator)")
        per_op_stats.to_excel(writer, sheet_name="Load_Balance", startrow=row+1, index=False)

        row = row + len(per_op_stats) + 4
        ws.write(row, 0, "Global Average Ops per Weekday")
        weekday_load[["Label","Avg_ops_per_night"]].to_excel(writer, sheet_name="Load_Balance", startrow=row+1, index=False)

        row = row + len(weekday_load) + 4
        ws.write(row, 0, "Per-Team by Weekday — Operator-Normalized Work Nights (sum of Ops/TeamSize)")
        per_team_weekday_ops_norm.to_excel(writer, sheet_name="Load_Balance", startrow=row+1, index=False)

        row = row + len(per_team_weekday_ops_norm) + 4
        ws.write(row, 0, "Per-Team — Nights per Operator (annual normalized)")
        per_team_norm.to_excel(writer, sheet_name="Load_Balance", startrow=row+1, index=False)

        row = row + len(per_team_norm) + 4
        ws.write(row, 0, "Per-Team by Weekday — Active Nights (≥1 operator)")
        present_pivot.to_excel(writer, sheet_name="Load_Balance", startrow=row+1, index=False)

        row = row + len(present_pivot) + 4
        ws.write(row, 0, "Per-Team by Weekday — Full Team Nights (≥ operators_per_team)")
        full_pivot.to_excel(writer, sheet_name="Load_Balance", startrow=row+1, index=False)


        # --- Nuits non couvertes (0 opérateur) ---
        cover["Date"] = pd.to_datetime(cover["Date"], errors="raise")
        uncovered = cover.loc[cover["Operators_on_duty"] < 1, ["Date"]].copy()
        uncovered["Weekday"] = uncovered["Date"].dt.dayofweek.map({
            0: "Mon", 1: "Tue", 2: "Wed", 3: "Thu", 4: "Fri", 5: "Sat", 6: "Sun"
        })
        uncovered = uncovered.rename(columns={"Date": "Uncovered_Night"})

        # Alerte console (optionnel)
        if len(uncovered) > 0:
            print(f"[Alerte] {len(uncovered)} night(s) non covered (0 operator).")

        # --- Écriture dans Load_Balance ---
        row = row + len(per_op_stats) + 4   # continue après tes blocs précédents
        ws.write(row, 0, "Uncovered Nights (0 operator)")
        uncovered.to_excel(writer, sheet_name="Load_Balance", startrow=row+1, index=False)

        # (optionnel) auto-filter + largeur colonnes
        for ws_name in ["Schedule","Summary","Teams","Coverage","Parameters","Load_Balance"]:
            ws2 = writer.sheets[ws_name]
            ws2.autofilter(0, 0, 0, max(0, ws2.dim_colmax))
            for col in range(0, max(1, ws2.dim_colmax)+1):
                ws2.set_column(col, col, 16)

    # Messages d'avertissement utiles
    return warnings


def generate_schedule_v2(params_path: str, output_path: str):
    P = read_params(params_path)

    total_ops = int(P["total_operators"])
    required_per_night = int(P["operators_per_team"])
    days_per_week = int(P["days_per_week"])
    working_days_per_week = int(P["working_days_per_week"])
    rest_days_per_week = days_per_week - working_days_per_week
    vac_weeks = int(P["vacation_weeks_per_year"])
    vac_min_block = int(P["consecutive_vacation_weeks"])
    training_days = int(P["training_days"])
    sickness_days = int(P["sickness_days"])
    max_hours = float(P["max_working_hours"])
    hpd = float(P["hours_per_day"])
    year = int(P["Year"])
    use_available_capacity = bool(P.get("use_available_capacity", True))
    vacation_preferred_start_month = int(P.get("vacation_preferred_start_month", 6))
    vacation_preferred_end_month = int(P.get("vacation_preferred_end_month", 9))

    labels = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    team_name = "Night_Pool"
    operators = [f"OP{str(i + 1).zfill(3)}" for i in range(total_ops)]
    warnings: List[str] = []

    minimum_ops = min_operators_required(days_per_week, required_per_night, working_days_per_week)
    if total_ops < minimum_ops:
        warnings.append(
            f"WARNING: total_operators ({total_ops}) is lower than the theoretical minimum "
            f"({minimum_ops}) required to cover {required_per_night} operators/night over "
            f"{days_per_week} days with {working_days_per_week} working days/operator/week."
        )

    start = date(year, 1, 1)
    end = date(year, 12, 31)
    dates = [d for d in daterange(start, end)]
    dates_set = set(dates)
    total_weeks = date(year, 12, 28).isocalendar().week

    rest_patterns, rest_warnings = build_operator_rest_patterns(
        operators, days_per_week, working_days_per_week, required_per_night
    )
    warnings.extend(rest_warnings)

    all_seed = 2025 + year
    op_to_vac_blocks: Dict[str, List[Tuple[date, date]]] = {}
    op_to_vac_days: Dict[str, set[date]] = {}
    op_to_training_days: Dict[str, List[date]] = {}
    op_to_sickness_days: Dict[str, List[date]] = {}

    for idx, op in enumerate(operators):
        seed = all_seed + idx + 1
        vac_blocks = plan_vacations_for_operator(
            year,
            total_weeks,
            vac_weeks,
            vac_min_block,
            seed,
            vacation_preferred_start_month,
            vacation_preferred_end_month,
        )
        vac_days = {
            d
            for a, b in vac_blocks
            for d in daterange(a, b)
            if d in dates_set
        }
        op_to_vac_blocks[op] = vac_blocks
        op_to_vac_days[op] = vac_days

        rest_wd = set(rest_patterns[op])
        eligible_for_absence = [
            d for d in dates
            if d.weekday() not in rest_wd and d not in vac_days
        ]
        picked_training = pick_scattered_days_from_pool(eligible_for_absence, training_days, seed + 7)
        remaining_for_sickness = [d for d in eligible_for_absence if d not in set(picked_training)]
        picked_sickness = pick_scattered_days_from_pool(remaining_for_sickness, sickness_days, seed + 17)

        if len(picked_training) < training_days:
            warnings.append(
                f"WARNING: {op} has only {len(picked_training)} eligible training day(s) "
                f"instead of {training_days}."
            )
        if len(picked_sickness) < sickness_days:
            warnings.append(
                f"WARNING: {op} has only {len(picked_sickness)} eligible sickness day(s) "
                f"instead of {sickness_days}."
            )

        op_to_training_days[op] = picked_training
        op_to_sickness_days[op] = picked_sickness

    rows = []
    for op in operators:
        rest_wd = set(rest_patterns[op])
        vac_days = op_to_vac_days[op]
        training_set = set(op_to_training_days[op])
        sickness_set = set(op_to_sickness_days[op])
        for d in dates:
            wd = d.weekday()
            if d in vac_days:
                assignment = "Vacation"
            elif wd in rest_wd:
                assignment = "Weekend"
            elif d in training_set:
                assignment = "Training"
            elif d in sickness_set:
                assignment = "Sickness"
            else:
                assignment = "Available"
            rows.append({
                "Date": d,
                "Weekday": labels[wd],
                "WeekdayIdx": wd,
                "Team": team_name,
                "Operator": op,
                "Assignment": assignment,
                "Hours": 0.0,
            })

    schedule = pd.DataFrame(rows)
    schedule["Date"] = pd.to_datetime(schedule["Date"], errors="raise")

    work_counter = {op: 0 for op in operators}
    hours_by_op = {op: 0.0 for op in operators}
    available_days_by_op = (
        schedule[schedule["Assignment"].eq("Available")]
        .groupby("Operator")
        .size()
        .reindex(operators, fill_value=0)
        .to_dict()
    )
    cap_nights_by_op = {op: int(math.floor(max_hours / hpd)) for op in operators}
    target_work_nights = {
        op: min(int(available_days_by_op[op]), int(cap_nights_by_op[op]))
        for op in operators
    }
    uncovered_rows = []

    def assign_work(row_idx: int) -> None:
        op = schedule.at[row_idx, "Operator"]
        schedule.at[row_idx, "Assignment"] = "Work"
        schedule.at[row_idx, "Hours"] = hpd
        work_counter[op] += 1
        hours_by_op[op] += hpd

    def sort_candidates(row_indices: List[int]) -> List[int]:
        return sorted(
            row_indices,
            key=lambda idx: (
                work_counter[schedule.at[idx, "Operator"]],
                hours_by_op[schedule.at[idx, "Operator"]],
                schedule.at[idx, "Operator"],
            ),
        )

    for d in sorted(schedule["Date"].unique()):
        day_mask = schedule["Date"].eq(d)
        candidate_idx = list(schedule.index[day_mask & schedule["Assignment"].eq("Available")])
        eligible = []
        for row_idx in candidate_idx:
            op = schedule.at[row_idx, "Operator"]
            if hours_by_op[op] + hpd <= max_hours:
                eligible.append(row_idx)

        eligible_sorted = sort_candidates(eligible)
        picked = eligible_sorted[:required_per_night]
        for row_idx in picked:
            assign_work(row_idx)

        if len(picked) < required_per_night:
            ts = pd.Timestamp(d)
            uncovered_rows.append({
                "Uncovered_Night": ts,
                "Weekday": labels[ts.weekday()],
                "Operators_on_duty": len(picked),
                "Required_operators": required_per_night,
                "Missing_operators": required_per_night - len(picked),
            })
            continue

    if use_available_capacity:
        for op in sorted(operators, key=lambda item: (work_counter[item], hours_by_op[item], item)):
            remaining = target_work_nights[op] - work_counter[op]
            if remaining <= 0:
                continue
            candidate_idx = list(schedule.index[
                schedule["Operator"].eq(op) & schedule["Assignment"].eq("Available")
            ])
            if not candidate_idx:
                continue
            candidate_by_date = {
                schedule.at[row_idx, "Date"].date(): row_idx
                for row_idx in candidate_idx
            }
            picked_dates = pick_scattered_days_from_pool(
                list(candidate_by_date.keys()),
                remaining,
                seed=all_seed + 10000 + int(op.replace("OP", "")),
            )
            for picked_date in picked_dates:
                row_idx = candidate_by_date[picked_date]
                if hours_by_op[op] + hpd <= max_hours:
                    assign_work(row_idx)

    if uncovered_rows:
        print(f"[Alerte] {len(uncovered_rows)} night(s) below required coverage.")
        warnings.append(
            f"WARNING: {len(uncovered_rows)} night(s) below required coverage "
            f"({required_per_night} operator(s) required)."
        )

    work_schedule = schedule[schedule["Assignment"] == "Work"]
    all_op_index = pd.MultiIndex.from_product([[team_name], operators], names=["Team", "Operator"])

    per_op = (
        work_schedule.groupby(["Team", "Operator"]).size()
        .reindex(all_op_index, fill_value=0)
        .rename("Work_nights")
        .reset_index()
    )
    per_team = (
        work_schedule.groupby("Team").size()
        .reindex([team_name], fill_value=0)
        .rename("Work_nights")
        .reset_index()
    )

    per_team_weekday = (
        work_schedule.groupby(["Team", "WeekdayIdx"]).size()
        .rename("Work_nights")
        .reset_index()
    )
    per_team_weekday_pivot = per_team_weekday.pivot_table(
        index="Team", columns="WeekdayIdx", values="Work_nights", aggfunc="sum", fill_value=0
    )
    per_team_weekday_pivot = per_team_weekday_pivot.reindex(index=[team_name], columns=range(7), fill_value=0)
    per_team_weekday_pivot = per_team_weekday_pivot.rename(columns={i: labels[i] for i in range(7)}).reset_index()

    per_op_stats = per_op.groupby("Team")["Work_nights"].agg(
        mean="mean", std="std", min="min", max="max", sum="sum", count="count"
    ).reset_index().rename(columns={"count": "Operators_in_pool"})

    cover = (
        work_schedule.groupby("Date").size()
        .reindex(sorted(schedule["Date"].unique()), fill_value=0)
        .rename("Operators_on_duty")
        .reset_index()
    )
    cover["Date"] = pd.to_datetime(cover["Date"], errors="raise")
    cover["Weekday"] = cover["Date"].dt.dayofweek.map({i: labels[i] for i in range(7)})
    cover["Required_operators"] = required_per_night
    cover["Coverage_gap"] = cover["Operators_on_duty"] - required_per_night
    cover["Status"] = np.where(cover["Coverage_gap"] >= 0, "OK", "UNDER_COVERED")
    cover = cover[["Date", "Weekday", "Operators_on_duty", "Required_operators", "Coverage_gap", "Status"]]

    daily_assignments = cover[["Date", "Weekday", "Operators_on_duty", "Required_operators", "Status"]].copy()
    assigned_by_date = (
        work_schedule.sort_values(["Date", "Operator"])
        .groupby("Date")["Operator"]
        .apply(list)
        .to_dict()
    )
    max_slots = max(required_per_night, int(cover["Operators_on_duty"].max()) if not cover.empty else 0)
    for i in range(max_slots):
        daily_assignments[f"Operator_{i + 1}"] = daily_assignments["Date"].map(
            lambda d, slot=i: assigned_by_date.get(d, [])[slot]
            if slot < len(assigned_by_date.get(d, []))
            else ""
        )
    work_only = work_schedule.sort_values(["Date", "Operator"])[
        ["Date", "Weekday", "Team", "Operator", "Assignment", "Hours"]
    ].copy()

    weekday_load = (
        cover.groupby("Weekday", sort=False)["Operators_on_duty"]
        .mean()
        .reindex(labels)
        .rename("Avg_ops_per_night")
        .reset_index()
        .rename(columns={"Weekday": "Label"})
    )
    weekday_load["Avg_ops_per_night"] = weekday_load["Avg_ops_per_night"].round(2)

    team_sizes = pd.DataFrame([{"Team": team_name, "Team_size": total_ops}])
    all_team_dates = pd.MultiIndex.from_product(
        [sorted(schedule["Date"].unique()), [team_name]], names=["Date", "Team"]
    ).to_frame(index=False)
    per_team_date_ops = (
        work_schedule.groupby(["Date", "Team"]).size()
        .rename("Ops")
        .reset_index()
    )
    per_team_date_ops = all_team_dates.merge(per_team_date_ops, on=["Date", "Team"], how="left").fillna({"Ops": 0})
    per_team_date_ops["Ops"] = per_team_date_ops["Ops"].astype(int)
    per_team_date_ops = per_team_date_ops.merge(team_sizes, on="Team", how="left")
    per_team_date_ops["WeekdayIdx"] = pd.to_datetime(per_team_date_ops["Date"]).dt.weekday
    per_team_date_ops["Ops_norm"] = per_team_date_ops["Ops"] / per_team_date_ops["Team_size"]
    per_team_date_ops["Team_present"] = per_team_date_ops["Ops"] >= 1
    per_team_date_ops["Team_full"] = per_team_date_ops["Ops"] >= required_per_night

    per_team_weekday_ops_norm = per_team_date_ops.pivot_table(
        index="Team", columns="WeekdayIdx", values="Ops_norm", aggfunc="sum", fill_value=0
    ).reindex(index=[team_name], columns=range(7), fill_value=0)
    per_team_weekday_ops_norm = per_team_weekday_ops_norm.rename(columns={i: labels[i] for i in range(7)}).reset_index()

    per_team_norm = per_team.merge(team_sizes, on="Team", how="left")
    per_team_norm["Nights_per_operator"] = (per_team_norm["Work_nights"] / per_team_norm["Team_size"]).round(2)

    present_pivot = per_team_date_ops.pivot_table(
        index="Team", columns="WeekdayIdx", values="Team_present", aggfunc="sum", fill_value=0
    ).reindex(index=[team_name], columns=range(7), fill_value=0)
    present_pivot = present_pivot.rename(columns={i: labels[i] for i in range(7)}).reset_index()

    full_pivot = per_team_date_ops.pivot_table(
        index="Team", columns="WeekdayIdx", values="Team_full", aggfunc="sum", fill_value=0
    ).reindex(index=[team_name], columns=range(7), fill_value=0)
    full_pivot = full_pivot.rename(columns={i: labels[i] for i in range(7)}).reset_index()

    piv = schedule.pivot_table(
        index=["Operator", "Team"], columns="Assignment", values="Date", aggfunc="count", fill_value=0
    )
    for col in ["Work", "Available", "Weekend", "Vacation", "Training", "Sickness"]:
        if col not in piv.columns:
            piv[col] = 0
    piv = piv.reset_index()
    hours_work = (
        work_schedule.groupby(["Operator", "Team"])["Hours"].sum()
        .reindex(piv.set_index(["Operator", "Team"]).index, fill_value=0)
        .values
    )
    piv["Hours_worked"] = hours_work
    piv["Hours_cap"] = max_hours
    piv["Cap_margin"] = piv["Hours_cap"] - piv["Hours_worked"]

    imbalance = per_op_stats["std"].fillna(0)
    if (imbalance > 2).any():
        print("[Alerte] Imbalance detected in the night pool (std dev > 2 nights).")

    teams_df = pd.DataFrame([{
        "Team": team_name,
        "Team_size": total_ops,
        "Required_operators_per_night": required_per_night,
        "Working_days_per_week": working_days_per_week,
        "Weekend_days_per_operator": rest_days_per_week,
        "Members": ", ".join(operators),
    }])
    patterns_df = pd.DataFrame([{
        "Operator": op,
        "Team": team_name,
        "Weekend_days": ", ".join(labels[d] for d in rest_patterns[op]),
        "Weekend_days_count": len(rest_patterns[op]),
    } for op in operators])
    preferred_start, preferred_end = month_window(
        year, vacation_preferred_start_month, vacation_preferred_end_month
    )
    vacation_rows = []
    for op in operators:
        for block_idx, (block_start, block_end) in enumerate(op_to_vac_blocks[op], start=1):
            vacation_rows.append({
                "Operator": op,
                "Block": block_idx,
                "Vacation_start": block_start,
                "Vacation_end": block_end,
                "Weeks": ((block_end - block_start).days + 1) // 7,
                "In_preferred_period": block_start >= preferred_start and block_end <= preferred_end,
            })
    vacation_blocks_df = pd.DataFrame(vacation_rows)
    params_df = pd.DataFrame([{"Param": k, "Value": v} for k, v in P.items()])
    uncovered = pd.DataFrame(uncovered_rows)
    if uncovered.empty:
        uncovered = pd.DataFrame(columns=[
            "Uncovered_Night", "Weekday", "Operators_on_duty",
            "Required_operators", "Missing_operators"
        ])

    try:
        import xlsxwriter  # noqa: F401
        excel_engine = "xlsxwriter"
    except Exception:
        excel_engine = "openpyxl"

    with pd.ExcelWriter(output_path, engine=excel_engine, datetime_format="yyyy-mm-dd") as writer:
        schedule.sort_values(["Date", "Operator"]).to_excel(writer, sheet_name="Schedule", index=False)
        daily_assignments.to_excel(writer, sheet_name="Daily_Assignments", index=False)
        work_only.to_excel(writer, sheet_name="Work_Only", index=False)
        piv.sort_values(["Team", "Operator"]).to_excel(writer, sheet_name="Summary", index=False)
        teams_df.to_excel(writer, sheet_name="Teams", index=False)
        patterns_df.to_excel(writer, sheet_name="Weekend_Patterns", index=False)
        vacation_blocks_df.to_excel(writer, sheet_name="Vacation_Blocks", index=False)
        cover.to_excel(writer, sheet_name="Coverage", index=False)
        params_df.to_excel(writer, sheet_name="Parameters", index=False)

        if excel_engine == "xlsxwriter":
            ws = writer.book.add_worksheet("Load_Balance")
        else:
            ws = writer.book.create_sheet("Load_Balance")
        writer.sheets["Load_Balance"] = ws

        def write_label(sheet, zero_row: int, zero_col: int, value: str) -> None:
            if excel_engine == "xlsxwriter":
                sheet.write(zero_row, zero_col, value)
            else:
                sheet.cell(row=zero_row + 1, column=zero_col + 1, value=value)

        row = 0
        load_tables = [
            ("Per-Operator Work Nights", per_op.sort_values(["Team", "Operator"])),
            ("Per-Team Work Nights", per_team.sort_values("Team")),
            ("Per-Team by Weekday (Work Nights)", per_team_weekday_pivot),
            ("Intra-Pool Stats (Work Nights per Operator)", per_op_stats),
            ("Global Average Ops per Weekday", weekday_load[["Label", "Avg_ops_per_night"]]),
            ("Per-Team by Weekday - Operator-Normalized Work Nights (sum of Ops/TeamSize)", per_team_weekday_ops_norm),
            ("Per-Team - Nights per Operator (annual normalized)", per_team_norm),
            ("Per-Team by Weekday - Active Nights (>=1 operator)", present_pivot),
            ("Per-Team by Weekday - Required Coverage Nights (>= operators_per_team)", full_pivot),
            ("Under-Covered Nights", uncovered),
        ]
        for title, df_out in load_tables:
            write_label(ws, row, 0, title)
            row += 1
            df_out.to_excel(writer, sheet_name="Load_Balance", startrow=row, index=False)
            row += len(df_out) + 3

        if excel_engine == "xlsxwriter":
            for ws_name in [
                "Schedule", "Daily_Assignments", "Work_Only", "Summary",
                "Teams", "Weekend_Patterns", "Vacation_Blocks", "Coverage", "Parameters", "Load_Balance"
            ]:
                ws2 = writer.sheets[ws_name]
                ws2.autofilter(0, 0, 0, max(0, ws2.dim_colmax))
                for col in range(0, max(1, ws2.dim_colmax) + 1):
                    ws2.set_column(col, col, 18)
        else:
            from openpyxl.utils import get_column_letter
            for ws_name in [
                "Schedule", "Daily_Assignments", "Work_Only", "Summary",
                "Teams", "Weekend_Patterns", "Vacation_Blocks", "Coverage", "Parameters", "Load_Balance"
            ]:
                ws2 = writer.sheets[ws_name]
                ws2.auto_filter.ref = ws2.dimensions
                for col in range(1, ws2.max_column + 1):
                    ws2.column_dimensions[get_column_letter(col)].width = 18

    return warnings


generate_schedule_legacy = generate_schedule
generate_schedule = generate_schedule_v2


# ---------------------- CLI ----------------------
def main(input_file: str | None = None,
         output_file: str | None = None,
         *,
         gui: bool = False) -> tuple[list[str], str]:
    """
    Point d'entrée réutilisable.
    - input_file/output_file : chemins explicites ; si None, on peut ouvrir des boîtes de dialogue si gui=True.
    - gui : True pour forcer les boîtes de dialogue quand des chemins manquent.
    Retourne (warnings, output_file).
    """
    # Faut-il ouvrir des boîtes ?
    need_gui = gui or not input_file or not output_file

    # Créer un root Tk temporaire uniquement si nécessaire et disponible
    root = None
    if need_gui:
        if Tk is None or filedialog is None:
            raise RuntimeError("GUI is unavailable (Tkinter). Please provide --input and --output.")
        root = Tk()
        root.withdraw()
        try:
            if not input_file:
                input_file = filedialog.askopenfilename(
                    title="Select the Parametres.xlsx file",
                    filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
                )
            if not output_file:
                output_file = filedialog.asksaveasfilename(
                    title="Save the schedule",
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")],
                    initialfile="planning.xlsx",
                )
        finally:
            # on détruit le root même si l'utilisateur annule
            try:
                root.destroy()
            except Exception:
                pass

    if not input_file:
        raise RuntimeError("No input file provided/selected.")
    if not output_file:
        raise RuntimeError("No output file provided/selected.")

    warnings = generate_schedule_v2(input_file, output_file)
    return warnings, output_file


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(
        description="Night operator scheduling with staggered weekends."
    )
    ap.add_argument("--input", help="Path to the Parametres.xlsx file (otherwise, dialog if --gui)", required=False)
    ap.add_argument("--output", help="Path to the output .xlsx file (otherwise, dialog if --gui)", required=False)
    ap.add_argument("--gui", action="store_true", help="Use file dialogs to select files")
    args = ap.parse_args()

    try:
        warns, outp = main(args.input, args.output, gui=args.gui)
        if warns:
            print("\n".join(str(w) for w in warns))
        print(f"Schedule generated -> {outp}")
    except Exception as e:
        # message clair en CLI
        print(f"Error : {e}")
        raise
