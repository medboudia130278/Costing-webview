"""
Vref — Adds a new column "PC filtered" for Recap rows

- Output columns: Field | Subsystem | Currency | Type | PC filtered | <Project columns>
- For Recap sheets:
    * PC_per_Subsystem  -> PC filtered = "Per subsystem"
    * PC_per_Type       -> PC filtered = "Per type"
    * PC_per_Period     -> PC filtered = "Per period"
- Non-Recap rows keep PC filtered empty.

Also includes all features (Subcontracting before Total Global Cost, merging identical Items, etc.).
"""

from __future__ import annotations
import sys
import re
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook

# ---- Constants ----
GENERAL_SHEET = "General Parameters"
SYNTHESIS_SHEET = "Synthesis"

# Recap via dedicated sheets
RECAP_SHEETS_PREFIXES = ["PC_per_Subsystem", "PC_per_Type", "PC_per_Period"]
RECAP_ITEM_COL = "Item"
RECAP_VALUE_COL = "PC wo kip/Firming"

KEY_PROJECT_TYPE = "Project_type"
KEY_PROJECT_NAME = "Project_name"
KEY_MAX_DAY = "max_day_technicians"
KEY_MAX_NIGHT = "max_night_technicians"

# Global fields (rows) — keep your custom additions here if needed
FIELDS = [
    "Region",
    "Country",
    "Project_context",
    "Bid_year",
    "Service_Year",
    "contract_duration_years",
    "Number_of_traction_substation",
    "Number_of_auxiliary_substation",
    "Number_of_MV_substation",
    "Number_of_trains",
    "Number_of_cars_per_train",
    "Mileage_per_train_per_year",
    "EMGT",
    "Traction_Voltage",
    "Rail_type",
    "Number_of_station",
    "walls_per_station",
    "APSD_per_wall",
    "L_total_single_track",
    "depot_total_single_track",
    "type_track_installation",
    "depot_track_installation",
    "feeding_system",
    "switch",
    "switch_depot",
    "diamond_crossing",
    "crossover",
    "point_machine",
    "recovery_technicians",
]

# Synthesis metrics
COST_FIELDS = ["Yearly Reparable Cost", "Yearly Total Cost"]
GLOBAL_COST_FIELD = "Total Global Cost"  # Only Overhaul/Renewal
SUBCON_FIELD = "Yearly Cost (Subcontracting)"

# ---- Helpers ----
def normalize_sheet_name(name: str) -> str:
    name = re.sub(r'[:\\/?*\[\]]', "_", str(name))
    return name[:31] if len(name) > 31 else name

def read_general_params(xlsx_path: Path) -> Optional[pd.DataFrame]:
    try:
        df = pd.read_excel(xlsx_path, sheet_name=GENERAL_SHEET, engine="openpyxl")
    except Exception as e:
        print(f"[WARN] {xlsx_path.name}: unable to read sheet '{GENERAL_SHEET}': {e}", file=sys.stderr)
        return None
    col_map = {str(c).strip().lower(): c for c in df.columns}
    nom_key    = 'nom'    if 'nom'    in col_map else ('name'  if 'name'  in col_map else None)
    valeur_key = 'valeur' if 'valeur' in col_map else ('value' if 'value' in col_map else None)
    if nom_key is None or valeur_key is None:
        print(f"[WARN] {xlsx_path.name}: sheet '{GENERAL_SHEET}' missing required columns 'Nom'/'Name' and 'Valeur'/'Value'", file=sys.stderr)
        return None
    df = df.rename(columns={col_map.get(nom_key): "Nom", col_map.get(valeur_key): "Valeur"})
    if "Subsystem" not in df.columns and "subsystem" in col_map:
        df = df.rename(columns={col_map["subsystem"]: "Subsystem"})
    if "Comments" not in df.columns and "comments" in col_map:
        df = df.rename(columns={col_map["comments"]: "Comments"})
    df = df[~df["Nom"].isna()].copy()
    df["Nom"] = df["Nom"].astype(str).str.strip()
    return df

def value_for_key(df: pd.DataFrame, key: str) -> Optional[str]:
    exact = df.loc[df["Nom"] == key, "Valeur"]
    if not exact.empty:
        return str(exact.iloc[0])
    ci = df.loc[df["Nom"].str.lower() == key.lower(), "Valeur"]
    if not ci.empty:
        return str(ci.iloc[0])
    return None

def read_synthesis(xlsx_path: Path) -> Optional[pd.DataFrame]:
    try:
        df = pd.read_excel(xlsx_path, sheet_name=SYNTHESIS_SHEET, engine="openpyxl")
    except Exception as e:
        print(f"[WARN] {xlsx_path.name}: unable to read sheet '{SYNTHESIS_SHEET}': {e}", file=sys.stderr)
        return None

    col_map = {str(c).strip().lower(): c for c in df.columns}
    required = {
        "subsystem": None,
        "currency": None,
        "type": None,
        "yearly reparable cost": None,
        "yearly total cost": None,
        "total global cost": None,
        "yearly cost (subcontracting)": None,
    }
    for k in list(required.keys()):
        if k in col_map:
            required[k] = col_map[k]

    def _flex(name: str) -> str:
        return name.replace(" ", "").replace("_", "").lower()

    if any(v is None for v in required.values()):
        flex_map = {_flex(k): v for k, v in col_map.items()}
        for key in list(required.keys()):
            if required[key] is None:
                cand = flex_map.get(_flex(key))
                if cand:
                    required[key] = cand

    if any(v is None for v in [required["subsystem"], required["currency"], required["type"]]):
        print(f"[WARN] {xlsx_path.name}: '{SYNTHESIS_SHEET}' missing key columns (Subsystem/Currency/Type)", file=sys.stderr)
        return None

    cols = [required["subsystem"], required["currency"], required["type"]]
    for k in ["yearly reparable cost", "yearly total cost", "total global cost", "yearly cost (subcontracting)"]:
        if required.get(k):
            cols.append(required[k])

    df = df[cols].copy()

    rename_map = {
        required["subsystem"]: "Subsystem",
        required["currency"]: "Currency",
        required["type"]: "Type",
    }
    if required.get("yearly reparable cost"):
        rename_map[required["yearly reparable cost"]] = "Yearly Reparable Cost"
    if required.get("yearly total cost"):
        rename_map[required["yearly total cost"]] = "Yearly Total Cost"
    if required.get("total global cost"):
        rename_map[required["total global cost"]] = "Total Global Cost"
    if required.get("yearly cost (subcontracting)"):
        rename_map[required["yearly cost (subcontracting)"]] = "Yearly Cost (Subcontracting)"

    df.rename(columns=rename_map, inplace=True)

    for c in ["Yearly Reparable Cost", "Yearly Total Cost", "Total Global Cost", "Yearly Cost (Subcontracting)"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    df = df[df["Subsystem"].notna() & df["Currency"].notna()].copy()
    return df

# ---------- Recap via dedicated sheets ----------

def read_recap_sheets(xlsx_path: Path) -> Dict[str, pd.DataFrame]:
    """
    Lit d'abord les feuilles dédiées (PC_per_*). Retourne un dict par *base*:
      {'PC_per_Subsystem': df, 'PC_per_Type': df, 'PC_per_Period': df} – uniquement celles trouvées/ayant des données.
    Les df retournés ont exactement 2 colonnes: ['Item', 'PC wo kip/Firming'] avec parsing numérique robuste (FR/EN),
    et on ignore la colonne en pourcentage si elle existe.
    """
    import re as _re

    def norm(s: str) -> str:
        return _re.sub(r'[^0-9a-zA-Z]+', '', str(s)).lower()

    BASES = ["PC_per_Subsystem", "PC_per_Type", "PC_per_Period"]
    wanted_norm = [norm(w) for w in BASES]

    # regex exacte (espaces optionnels) pour éviter la colonne %PC...
    target_regex = _re.compile(r"^pc\s*wo\s*kip\s*/\s*firming$", _re.IGNORECASE)

    out: Dict[str, pd.DataFrame] = {}
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        sheetnames = list(wb.sheetnames)
    except Exception as e:
        print(f"[WARN] {xlsx_path.name}: cannot open workbook: {e}", file=sys.stderr)
        return out
    finally:
        try:
            wb.close()
        except Exception:
            pass

    def base_from_sheetname(s: str) -> Optional[str]:
        sn = norm(s)
        for base, w in zip(BASES, wanted_norm):
            if sn.startswith(w):
                return base
        return None

    # 1) tenter les feuilles dédiées
    for sname in sheetnames:
        base = base_from_sheetname(sname)
        if not base:
            continue
        try:
            df0 = pd.read_excel(xlsx_path, sheet_name=sname, engine="openpyxl")
        except Exception as e:
            print(f"[WARN] {xlsx_path.name}: cannot read sheet '{sname}': {e}", file=sys.stderr)
            continue
        if df0 is None or df0.empty:
            continue

        cols = [c for c in df0.columns if isinstance(c, str)]

        # Item
        def match_item(col: str) -> bool:
            c = col.strip().lower()
            return c == "item" or c.replace(" ", "").replace("_", "") == "item"
        item_candidates = [c for c in cols if match_item(c)]
        item_col = item_candidates[0] if item_candidates else None

        # Value (préférer non-%)
        non_percent = [c for c in cols if '%' not in c]
        exact_candidates = [c for c in non_percent if target_regex.match(c.strip().lower())]

        def norm_str(s: str) -> str:
            return re.sub(r'[^0-9a-zA-Z]+', '', s).lower()
        norm_target = norm_str(RECAP_VALUE_COL)
        norm_candidates = [c for c in non_percent if norm_str(c) == norm_target]

        value_col = None
        if exact_candidates:
            value_col = exact_candidates[0]
        elif norm_candidates:
            # le plus proche en longueur de l'étiquette cible
            value_col = min(norm_candidates, key=lambda s: abs(len(s) - len(RECAP_VALUE_COL)))
        else:
            # 3) Fallback "contains" basé sur la version normalisée
            partial_candidates = [c for c in non_percent if norm_target in norm_str(c)]
            if partial_candidates:
                value_col = min(partial_candidates, key=lambda s: abs(len(s) - len(RECAP_VALUE_COL)))
            else:
                # 4) NOUVEAU : toute colonne qui commence par "PC " (PC + espace), en évitant les colonnes en %
                pc_space_regex = re.compile(r'^\s*pc\s+', re.IGNORECASE)  # autorise espaces initiaux, exige un espace après PC
                for cname in list(df0.columns):
                    if isinstance(cname, str) and ('%' not in cname) and pc_space_regex.match(cname):
                        value_col = cname
                        break

        if not item_col or not value_col:
            # feuille dédiée présente mais pas les bonnes colonnes -> on laisse le fallback gérer
            continue

        sub = df0[[item_col, value_col]].copy()
        sub.columns = [RECAP_ITEM_COL, RECAP_VALUE_COL]

        def _parse_num(x):
            if pd.isna(x):
                return None
            s = str(x).strip().replace("\u00A0", "").replace(" ", "")
            if s == "":
                return None
            if "," in s and "." in s:
                s = s.replace(".", "").replace(",", ".")
            elif "," in s:
                s = s.replace(",", ".")
            s = re.sub(r"[^0-9\.\-]", "", s)
            try:
                return float(s) if s not in ("", ".", "-", "-.") else None
            except Exception:
                return None

        sub[RECAP_VALUE_COL] = sub[RECAP_VALUE_COL].apply(_parse_num)

        # garder seulement si non vide après nettoyage
        if sub.dropna(how="all").empty:
            continue
        out[base] = sub

    return out


def read_recap_structured_tables(xlsx_path: Path) -> Dict[str, pd.DataFrame]:
    """
    Fallback: lit les tables structurées PC_per_* sur la feuille 'Recap'.
    Compatible avec différentes versions d'openpyxl : récupère les vrais objets Table,
    même si ws.tables ne donne que des noms (strings).
    """
    import re as _re

    BASES = ["PC_per_Subsystem", "PC_per_Type", "PC_per_Period"]

    def norm(s: str) -> str:
        return _re.sub(r'[^0-9a-zA-Z]+', '', str(s)).lower()

    try:
        # IMPORTANT: read_only=False pour accéder aux tables structurées
        wb = load_workbook(xlsx_path, read_only=False, data_only=True)
    except Exception as e:
        print(f"[WARN] {xlsx_path.name}: cannot open workbook for structured tables: {e}", file=sys.stderr)
        return {}

    # trouver la sheet Recap (case-insensitive)
    ws = None
    if "Recap" in wb.sheetnames:
        ws = wb["Recap"]
    else:
        for s in wb.sheetnames:
            if str(s).strip().lower() == "recap":
                ws = wb[s]
                break
    if ws is None:
        wb.close()
        return {}

    # Construire un dict {displayName: TableObj} de manière robuste
    tables_dict = {}
    # 1) si ws._tables existe, on l'utilise (liste d'objets Table)
    if hasattr(ws, "_tables") and ws._tables:
        for tbl in ws._tables:
            try:
                tables_dict[str(tbl.displayName)] = tbl
            except Exception:
                pass

    # 2) si ws.tables est un dict/iterable, on tente d'y retrouver les objets via _tables
    try:
        raw_tables = ws.tables  # peut être dict-like ou autre selon versions
    except Exception:
        raw_tables = {}

    if raw_tables:
        # raw_tables pourrait être un dict {name: <unknown>} ou {name: Table}
        # Harmonisation: pour chaque nom, chercher l'objet dans ws._tables
        try:
            iterable = getattr(raw_tables, "items", None)
            if callable(iterable):
                names = [k for k, _ in raw_tables.items()]
            else:
                # si ce n'est pas un dict, essayons de le traiter comme iterable de noms
                names = list(raw_tables)  # peut lever, d'où le try/except global
        except Exception:
            names = []

        for name in names:
            name_str = str(name)
            if name_str not in tables_dict and hasattr(ws, "_tables"):
                for tbl in ws._tables:
                    if str(getattr(tbl, "displayName", "")) == name_str:
                        tables_dict[name_str] = tbl
                        break
            # si raw_tables contient déjà un objet avec .ref, on peut aussi le prendre
            val = raw_tables[name] if hasattr(raw_tables, "__getitem__") else None
            if name_str not in tables_dict and hasattr(val, "ref"):
                tables_dict[name_str] = val

    # Helper: base depuis le nom de la table
    def base_from_tablename(tname: str) -> Optional[str]:
        tn = norm(tname)
        for base in BASES:
            if tn.startswith(norm(base)):
                return base
        return None

    target_regex = _re.compile(r"^pc\s*wo\s*kip\s*/\s*firming$", _re.IGNORECASE)

    out: Dict[str, pd.DataFrame] = {}

    for tname, tbl in tables_dict.items():
        base = base_from_tablename(tname)
        if not base:
            continue

        # Sécuriser l'accès à la ref
        ref = getattr(tbl, "ref", None)
        if not isinstance(ref, str):
            # dernier recours: on saute cette table
            continue

        try:
            cells = ws[ref]  # ex "B2:G20"
        except Exception:
            continue

        rows = [[c.value for c in row] for row in cells]
        if not rows:
            continue
        header, *body = rows
        df0 = pd.DataFrame(body, columns=header)
        if df0 is None or df0.empty:
            continue

        cols = [c for c in df0.columns if isinstance(c, str)]

        # --- Item column ---
        def match_item(col: str) -> bool:
            c = col.strip().lower()
            return c == "item" or c.replace(" ", "").replace("_", "") == "item"
        item_candidates = [c for c in cols if match_item(c)]
        item_col = item_candidates[0] if item_candidates else None

        # --- Value column: prefer non-percent
        non_percent = [c for c in cols if '%' not in c]
        exact_candidates = [c for c in non_percent if target_regex.match(c.strip().lower())]

        def norm_str(s: str) -> str:
            return _re.sub(r'[^0-9a-zA-Z]+', '', s).lower()
        norm_target = norm_str(RECAP_VALUE_COL)
        norm_candidates = [c for c in non_percent if norm_str(c) == norm_target]

        value_col = None
        if exact_candidates:
            value_col = exact_candidates[0]
        elif norm_candidates:
            # le plus proche en longueur de l'étiquette cible
            value_col = min(norm_candidates, key=lambda s: abs(len(s) - len(RECAP_VALUE_COL)))
        else:
            # 3) Fallback "contains" basé sur la version normalisée
            partial_candidates = [c for c in non_percent if norm_target in norm_str(c)]
            if partial_candidates:
                value_col = min(partial_candidates, key=lambda s: abs(len(s) - len(RECAP_VALUE_COL)))
            else:
                # 4) NOUVEAU : toute colonne qui commence par "PC " (PC + espace), en évitant les colonnes en %
                pc_space_regex = re.compile(r'^\s*pc\s+', re.IGNORECASE)  # autorise espaces initiaux, exige un espace après PC
                for cname in list(df0.columns):
                    if isinstance(cname, str) and ('%' not in cname) and pc_space_regex.match(cname):
                        value_col = cname
                        break

        if not item_col or not value_col:
            print(f"[WARN] {xlsx_path.name}: structured table '{tname}' missing expected columns ('{RECAP_ITEM_COL}','{RECAP_VALUE_COL}'). Found: {list(df0.columns)}")
            continue

        sub = df0[[item_col, value_col]].copy()
        sub.columns = [RECAP_ITEM_COL, RECAP_VALUE_COL]

        def _parse_num(x):
            if pd.isna(x):
                return None
            s = str(x).strip().replace("\u00A0", "").replace(" ", "")
            if s == "":
                return None
            if "," in s and "." in s:
                s = s.replace(".", "").replace(",", ".")
            elif "," in s:
                s = s.replace(",", ".")
            s = _re.sub(r"[^0-9\.\-]", "", s)
            try:
                return float(s) if s not in ("", ".", "-", "-.") else None
            except Exception:
                return None

        sub[RECAP_VALUE_COL] = sub[RECAP_VALUE_COL].apply(_parse_num)
        if not sub.dropna(how="all").empty:
            out[base] = sub

    wb.close()
    return out


# ---------- Main pipeline ----------

def process_file(xlsx_path: Path):
    gdf = read_general_params(xlsx_path)
    if gdf is None:
        return None

    project_type = value_for_key(gdf, KEY_PROJECT_TYPE)
    project_name = value_for_key(gdf, KEY_PROJECT_NAME)
    if not project_type:
        print(f"[WARN] {xlsx_path.name}: '{KEY_PROJECT_TYPE}' not found", file=sys.stderr)
        return None
    if not project_name:
        project_name = xlsx_path.stem

    globals_dict = {field: value_for_key(gdf, field) for field in FIELDS}

    if "Subsystem" in gdf.columns:
        sub_df = gdf[gdf["Subsystem"].notna()].copy()
    else:
        sub_df = pd.DataFrame(columns=["Nom","Valeur","Subsystem"])

    sub_df["_nom_lower"] = sub_df["Nom"].str.lower()
    max_day = sub_df[sub_df["_nom_lower"] == KEY_MAX_DAY]
    max_night = sub_df[sub_df["_nom_lower"] == KEY_MAX_NIGHT]

    def _prep(df_in: pd.DataFrame, colname: str) -> pd.DataFrame:
        if df_in.empty:
            return pd.DataFrame(columns=["Subsystem", colname]).set_index("Subsystem")
        tmp = df_in[["Subsystem","Valeur"]].copy()
        tmp["_num"] = pd.to_numeric(tmp["Valeur"], errors="coerce")
        agg = tmp.groupby("Subsystem").agg(**{colname: ("_num", "max")})
        fallback_needed = agg[colname].isna()
        if fallback_needed.any():
            first_vals = tmp.groupby("Subsystem")["Valeur"].first()
            agg.loc[fallback_needed, colname] = first_vals.loc[fallback_needed].values
        return agg

    day_df = _prep(max_day, "Day")
    night_df = _prep(max_night, "Night")
    maxtechs_df = day_df.join(night_df, how="outer").sort_index()

    synth_df = read_synthesis(xlsx_path)

    return project_type, project_name, globals_dict, maxtechs_df, synth_df

def build_recap_rows_for_file(xlsx_path: Path):
    gdf = read_general_params(xlsx_path)
    if gdf is None:
        return None
    project_type = value_for_key(gdf, KEY_PROJECT_TYPE)
    project_name = value_for_key(gdf, KEY_PROJECT_NAME)
    if not project_type:
        return None
    if not project_name:
        project_name = xlsx_path.stem

    # 1) D'abord: feuilles dédiées
    dedicated = read_recap_sheets(xlsx_path)  # dict base->df
    # 2) Fallback: tables structurées dans "Recap" (si base manquante ou vide)
    structured = read_recap_structured_tables(xlsx_path)  # dict base->df

    # fusion logique: prioriser la feuille dédiée si présente & non vide, sinon utiliser la table structurée
    merged: Dict[str, pd.DataFrame] = {}
    for base in ["PC_per_Subsystem", "PC_per_Type", "PC_per_Period"]:
        df_use = None
        if base in dedicated and not dedicated[base].empty:
            df_use = dedicated[base]
        elif base in structured and not structured[base].empty:
            df_use = structured[base]
        if df_use is not None:
            merged[base] = df_use

    if not merged:
        return project_type, project_name, []

    def label_for_base(base: str) -> str:
        b = base.lower()
        if b.startswith("pc_per_subsystem"):
            return "Per subsystem"
        if b.startswith("pc_per_type"):
            return "Per type"
        if b.startswith("pc_per_period"):
            return "Per period"
        return ""

    rows = []
    for base, df in merged.items():
        label = label_for_base(base)
        for _, r in df.iterrows():
            item = r.get(RECAP_ITEM_COL)
            val = r.get(RECAP_VALUE_COL)
            if pd.isna(item):
                continue
            rows.append({
                "Field": str(item),
                "Subsystem": "",
                "Currency": "EUR",    # EUR par défaut pour Recap
                "Type": "",
                "PC filtered": label,
                "_value": None if pd.isna(val) else val
            })

    return project_type, project_name, rows

def build_single_sheet_frames(benchmark_dir: Path, recap_dir: Path | None = None) -> Dict[str, pd.DataFrame]:
    per_type_projects: Dict[str, Dict[str, Dict[str, Optional[str]]]] = {}
    per_type_maxtechs: Dict[str, Dict[str, pd.DataFrame]] = {}
    per_type_costs: Dict[str, Dict[str, pd.DataFrame]] = {}

    xlsx_files = sorted([p for p in benchmark_dir.glob("*.xlsx") if p.is_file()])
    if not xlsx_files:
        print(f"[INFO] No .xlsx files found in '{benchmark_dir}'.", file=sys.stderr)

    for xlsx in xlsx_files:
        res = process_file(xlsx)
        if res is None:
            continue
        ptype, pname, globals_dict, maxtechs_df, costs_df = res
        per_type_projects.setdefault(ptype, {})[pname] = globals_dict
        per_type_maxtechs.setdefault(ptype, {})[pname] = maxtechs_df
        if costs_df is not None:
            per_type_costs.setdefault(ptype, {})[pname] = costs_df

    per_type_frames: Dict[str, pd.DataFrame] = {}
    for ptype, proj_map in per_type_projects.items():
        project_names = sorted(proj_map.keys())

        rows = []
        # Global fields
        for field in FIELDS:
            row = {"Field": field, "Subsystem": "", "Currency": "", "Type": "", "PC filtered": ""}
            for pname in project_names:
                row[pname] = proj_map[pname].get(field)
            rows.append(row)

        # Subsystem max techs
        all_subsystems = set()
        for df in per_type_maxtechs.get(ptype, {}).values():
            if df is not None and not df.empty:
                all_subsystems.update(df.index.tolist())
        for subsystem in sorted(all_subsystems):
            row_day = {"Field": KEY_MAX_DAY, "Subsystem": subsystem, "Currency": "", "Type": "", "PC filtered": ""}
            row_night = {"Field": KEY_MAX_NIGHT, "Subsystem": subsystem, "Currency": "", "Type": "", "PC filtered": ""}
            for pname in project_names:
                df = per_type_maxtechs.get(ptype, {}).get(pname)
                if df is not None and not df.empty and subsystem in df.index:
                    row_day[pname] = df.at[subsystem, "Day"]
                    row_night[pname] = df.at[subsystem, "Night"]
                else:
                    row_day[pname] = None
                    row_night[pname] = None
            rows.append(row_day)
            rows.append(row_night)

        # Yearly Reparable/Total by (Subsystem, Currency)
        all_pairs = set()
        for df in per_type_costs.get(ptype, {}).values():
            if df is not None and not df.empty:
                all_pairs.update(list(df[["Subsystem","Currency"]].drop_duplicates().itertuples(index=False, name=None)))
        for subsystem, currency in sorted(all_pairs, key=lambda x: (x[0], str(x[1]))):
            for field in COST_FIELDS:
                row_cost = {"Field": field, "Subsystem": subsystem, "Currency": currency if currency is not None else "", "Type": "", "PC filtered": ""}
                for pname in project_names:
                    df = per_type_costs.get(ptype, {}).get(pname)
                    if df is not None and not df.empty and field in df.columns:
                        match = df[(df["Subsystem"] == subsystem) & (df["Currency"] == currency)]
                        val = match[field].sum() if not match.empty else None
                        row_cost[pname] = val if pd.notna(val) else None
                    else:
                        row_cost[pname] = None
                rows.append(row_cost)

        # Subcontracting BEFORE Total Global Cost
        all_subcon = set()
        for df in per_type_costs.get(ptype, {}).values():
            if df is not None and not df.empty and "Type" in df.columns and SUBCON_FIELD in df.columns:
                excluded = {"preventive", "corrective", "overhaul", "renewal"}
                typ_norm = df["Type"].astype(str).str.strip().str.lower()
                mask = ~typ_norm.isin(excluded)
                _df = df[mask].copy()
                triples = _df[["Subsystem","Currency","Type"]].drop_duplicates()
                all_subcon.update(list(triples.itertuples(index=False, name=None)))
        for subsystem, currency, typ in sorted(all_subcon, key=lambda x: (x[0], str(x[1]), str(x[2]))):
            row_subc = {"Field": SUBCON_FIELD, "Subsystem": subsystem, "Currency": currency if currency is not None else "", "Type": typ if typ is not None else "", "PC filtered": ""}
            for pname in project_names:
                df = per_type_costs.get(ptype, {}).get(pname)
                if df is not None and not df.empty and SUBCON_FIELD in df.columns:
                    excluded = {"preventive", "corrective", "overhaul", "renewal"}
                    typ_norm = df["Type"].astype(str).str.strip().str.lower()
                    mask = ~typ_norm.isin(excluded)
                    match = df[(df["Subsystem"] == subsystem) & (df["Currency"] == currency) & (df["Type"] == typ) & mask]
                    val = match[SUBCON_FIELD].sum() if not match.empty else None
                    row_subc[pname] = val if pd.notna(val) else None
                else:
                    row_subc[pname] = None
            rows.append(row_subc)

        # Total Global Cost — only Overhaul/Renewal
        all_triples = set()
        for df in per_type_costs.get(ptype, {}).values():
            if df is not None and not df.empty and "Type" in df.columns:
                _df = df[df["Type"].astype(str).str.contains("Overhaul|Renewal", case=False, na=False)].copy()
                triples = _df[["Subsystem","Currency","Type"]].drop_duplicates()
                all_triples.update(list(triples.itertuples(index=False, name=None)))
        for subsystem, currency, typ in sorted(all_triples, key=lambda x: (x[0], str(x[1]), str(x[2]))):
            row_tgc = {"Field": GLOBAL_COST_FIELD, "Subsystem": subsystem, "Currency": currency if currency is not None else "", "Type": typ if typ is not None else "", "PC filtered": ""}
            for pname in project_names:
                df = per_type_costs.get(ptype, {}).get(pname)
                if df is not None and not df.empty and GLOBAL_COST_FIELD in df.columns:
                    match = df[(df["Subsystem"] == subsystem) & (df["Currency"] == currency) & (df["Type"] == typ)]
                    match = match[match["Type"].astype(str).str.contains("Overhaul|Renewal", case=False, na=False)]
                    val = match[GLOBAL_COST_FIELD].sum() if not match.empty else None
                    row_tgc[pname] = val if pd.notna(val) else None
                else:
                    row_tgc[pname] = None
            rows.append(row_tgc)

        df_out = pd.DataFrame(rows)
        df_out = df_out[["Field", "Subsystem", "Currency", "Type", "PC filtered"] + project_names]
        per_type_frames[ptype] = df_out

    # ---- Merge Recap-based rows (optional second folder) ----
    if recap_dir is not None:
        per_type_recap_rows: Dict[str, Dict[str, List[dict]]] = {}
        xlsx_files2 = sorted([p for p in recap_dir.glob("*.xlsx") if p.is_file()])
        for x2 in xlsx_files2:
            res = build_recap_rows_for_file(x2)
            if res is None:
                continue
            ptype2, pname2, rows2 = res
            if not rows2:
                continue
            per_type_recap_rows.setdefault(ptype2, {}).setdefault(pname2, []).extend(rows2)

        for ptype2, proj_map2 in per_type_recap_rows.items():
            if ptype2 not in per_type_frames:
                project_names = sorted(proj_map2.keys())
                df_empty = pd.DataFrame(columns=["Field","Subsystem","Currency","Type","PC filtered"] + project_names)
                per_type_frames[ptype2] = df_empty

            df_curr = per_type_frames[ptype2]
            specials = ["Field","Subsystem","Currency","Type","PC filtered"]
            existing_projects = [c for c in df_curr.columns if c not in specials]
            needed_projects = sorted(set(existing_projects) | set(proj_map2.keys()))
            if set(needed_projects) != set(existing_projects):
                new_projects = [p for p in needed_projects if p not in existing_projects]
                df_curr = df_curr.reindex(columns=specials + existing_projects + new_projects, fill_value=None)

            def norm_field(s: str) -> str:
                if s is None:
                    return ""
                return " ".join(str(s).strip().lower().split())

            recap_mask = (df_curr["Subsystem"].astype(str).str.strip() == "") & \
                         (df_curr["Type"].astype(str).str.strip() == "")
            field_to_idx = {}
            for idx in df_curr.index[recap_mask]:
                k = norm_field(df_curr.at[idx, "Field"])
                if k and k not in field_to_idx:
                    field_to_idx[k] = idx

            for pname2, rows_list in proj_map2.items():
                if pname2 not in df_curr.columns:
                    df_curr[pname2] = None

                for r in rows_list:
                    f_display = r.get("Field", "")
                    k = norm_field(f_display)
                    val = r.get("_value")
                    label = r.get("PC filtered", "")

                    if k in field_to_idx:
                        df_curr.at[field_to_idx[k], pname2] = val
                        # Keep existing PC filtered label if already there; if empty, set it
                        if not str(df_curr.at[field_to_idx[k], "PC filtered"]).strip() and label:
                            df_curr.at[field_to_idx[k], "PC filtered"] = label
                        if not str(df_curr.at[field_to_idx[k], "Currency"]).strip():
                            df_curr.at[field_to_idx[k], "Currency"] = "EUR"
                    else:
                        new_row = {col: None for col in df_curr.columns}
                        new_row["Field"] = f_display
                        new_row["Subsystem"] = ""
                        new_row["Currency"] = "EUR"
                        new_row["Type"] = ""
                        new_row["PC filtered"] = label
                        new_row[pname2] = val
                        df_curr = pd.concat([df_curr, pd.DataFrame([new_row])], ignore_index=True)
                        new_idx = df_curr.index[-1]
                        field_to_idx[k] = new_idx

            per_type_frames[ptype2] = df_curr

    return per_type_frames

def build_all_projects_frame(frames_by_type: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    Construit la feuille 'All_Projects' en fusionnant toutes les feuilles par type.
    - Colonnes fixes: Field, Subsystem, Currency, Type, PC filtered
    - Colonnes projets: union de tous les projets
    - Ajoute en 1ère ligne: Field='Project_type' et, pour chaque projet, sa valeur Project_type
    """
    specials = ["Field", "Subsystem", "Currency", "Type", "PC filtered"]

    # 1) Mapping projet -> project_type, déduit des frames existants
    project_type_of: Dict[str, str] = {}
    for ptype, df in frames_by_type.items():
        for col in df.columns:
            if col not in specials:
                project_type_of[col] = ptype  # ce projet appartient à ce type

    # 2) Collecte ordonnée des lignes (clé = tuple des 5 colonnes fixes)
    rows_map: Dict[tuple, dict] = {}
    order: List[tuple] = []

    for _, df in frames_by_type.items():
        for _, r in df.iterrows():
            key = (
                r.get("Field", ""),
                str(r.get("Subsystem", "")),
                str(r.get("Currency", "")),
                str(r.get("Type", "")),
                str(r.get("PC filtered", "")),
            )
            if key not in rows_map:
                base = {
                    "Field": key[0],
                    "Subsystem": key[1],
                    "Currency": key[2],
                    "Type": key[3],
                    "PC filtered": key[4],
                }
                rows_map[key] = base
                order.append(key)

            row = rows_map[key]
            # reporte toutes les valeurs de projet présentes dans ce df
            for col in df.columns:
                if col in specials:
                    continue
                val = r.get(col, None)
                if pd.notna(val):
                    row[col] = val
                else:
                    row.setdefault(col, None)

    # 3) Colonnes projets : tri par nom pour la stabilité
    project_names = sorted(project_type_of.keys())

    # 4) Construire les lignes dans l'ordre rencontré
    rows = []
    for key in order:
        row = rows_map[key]
        # garantir toutes les colonnes projets
        for p in project_names:
            row.setdefault(p, None)
        rows.append(row)

    # 5) Insérer la ligne "Project_type" tout en haut
    projtype_row = {"Field": "Project_type", "Subsystem": "", "Currency": "", "Type": "", "PC filtered": ""}
    for p in project_names:
        projtype_row[p] = project_type_of.get(p, "")
    rows.insert(0, projtype_row)

    # 6) DataFrame final
    df_all = pd.DataFrame(rows, columns=specials + project_names)
    return df_all


def save_single_sheet_per_type(frames_by_type: Dict[str, pd.DataFrame], output_path: Path) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for ptype, df in frames_by_type.items():
            sheet_name = normalize_sheet_name(ptype)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        df_all = build_all_projects_frame(frames_by_type)
        df_all.to_excel(writer, sheet_name="All_Projects", index=False)

def main() -> int:
    root = tk.Tk()
    root.withdraw()

    folder = filedialog.askdirectory(title="Select the folder containing the .xlsx files")
    if not folder:
        messagebox.showinfo("Cancelled", "No folder selected.")
        return 1
    benchmark_dir = Path(folder)
    if not benchmark_dir.exists() or not benchmark_dir.is_dir():
        messagebox.showerror("Error", f"Invalid folder: {benchmark_dir}")
        return 2

    use_second = messagebox.askyesno("Option", "Add a second folder containing 'Recap' files (dedicated sheets)?")
    recap_dir = None
    if use_second:
        folder2 = filedialog.askdirectory(title="Select the 'Recap' folder (with PC_per_* sheets)")
        if folder2:
            recap_dir = Path(folder2)
            if not recap_dir.exists() or not recap_dir.is_dir():
                messagebox.showerror("Error", f"Invalid folder: {recap_dir}")
                return 2

    frames_by_type = build_single_sheet_frames(benchmark_dir, recap_dir)
    if not frames_by_type:
        messagebox.showwarning("Warning", "No valid project found.")
        return 3

    output_file = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel Workbook", "*.xlsx")],
        title="Select the output file",
        initialfile="Benchmark_Summary_By_ProjectType.xlsx",
    )
    if not output_file:
        messagebox.showinfo("Cancelled", "No output file selected.")
        return 1

    try:
        save_single_sheet_per_type(frames_by_type, Path(output_file))
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save the file:\n{e}")
        return 4

    messagebox.showinfo("Completed", f"File successfully generated :\n{output_file}")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
